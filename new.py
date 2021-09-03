import numpy as np
import xgboost as xgb
import pandas as pd
from joblib import load 
import tkinter as tk
import copy
import tkinter.messagebox as mb
import sys
import xlwings as xw
from openpyxl.formatting.rule import IconSetRule
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import time
import pythoncom
#import threading
import traceback
import os

global filename
# raw data
ff = open("filename.txt","r")
filename = ff.read()
ff.close() 
         
def run():
    all_line = 467
    def change_schedule(now_schedule,all_schedule):
        entbar.coords(fill_rec, (0, 0, (now_schedule/all_schedule)*500, 36))
        root.update()
    def get_cur_info():
        try:
            raise Exception
        except:
            f = sys.exc_info()[2].tb_frame.f_back    
        return f.f_lineno

    # raw data
    ff_output = open("filename output.txt","r")
    file_output = ff_output.read()
    ff_output.close() 
    file_output_bfstd = file_output[:-5] + '_' + time.strftime("%Y%m%d") + '.xlsx'
    raw_data_SR = pd.read_excel(filename, sheet_name = 'Summary Report')
    raw_data_MR = pd.read_excel(filename, sheet_name = 'MLFB Report')
    # extra data
    index_data = pd.read_excel('PACI Price Index by PG.xlsx')
    index_data_extra = pd.read_excel('PACI Price Index.xlsx')
    fixed_data = pd.read_excel('20191202 PA SAP List Price Lookup.xlsx')
    c_value = pd.read_excel('SAP_EDI_Value - PACI FY20 (by CNOC).xlsx')
    c_value = c_value[['CNOC',
                       'segment_ADBA',
                       'sum_PurchaseVolume_201516',
                       'sum_PurchaseVolume_201617',
                       'sum_PurchaseVolume_201718',
                       'sum_PurchaseVolume_201819']]
    c_value['CNOC'] = c_value['CNOC'].astype(str)
    c_branch = pd.read_excel('FY20 Company List_20200408 (by Customer).xlsx').drop('Company Name(CN)', axis = 1)
    c_branch['CNOC'] = c_branch['CNOC'].astype(str)
    df_approvers = pd.read_excel('Approvers before Chen Xin Guang.xlsx')
    approvers = df_approvers['Approver'].str.upper().unique()

    # 4 model
    sc = load('std_scaler.bin')
    mean = xgb.Booster(model_file = 'mean.model')
    upper80 = load('upper_80.joblib')
    upper95 = load('upper_95.joblib')
  
    now_line = get_cur_info()
    change_schedule(now_line,all_line)
    try:
        dataMR_need = raw_data_MR[['SPR NO.', 'Price Group', 'Agreed Qty',
                               'Total Volume After Standard Discount Off', 'Gross Margin', 'Gross Margin After SGA',
                               'List Price', 'MLFB', 'Total Volume After Special Discount Off']]
        dataSR_need = raw_data_SR[['SPR NO.','Category', 'End User NO.', 'Sales Region', 'SPR Status',
                               'End User Name(CN)','Submitted Date','Fiscal Year',
                               'Name1','Cost1','Name2','Cost2','Name3',
                               'Cost3','Name4','Cost4','Name5','Cost5']]
    except Exception as e:
        mb.showerror(title="Error", message='Missing column(s)!\n\n' + repr(e))  
        root.destroy()

    dataSR_need['End User NO.'] = dataSR_need['End User NO.'].astype(str)
    
    # name&cost
    dataSR_need = dataSR_need[(dataSR_need['SPR Status']=='Submitted')&(dataSR_need['Category'].str.contains('CI'))].reset_index().drop(['index'], axis = 1)
    dataSR_need[['Name1', 'Name2', 'Name3', 'Name4', 'Name5']] = dataSR_need[['Name1', 'Name2', 'Name3', 'Name4', 'Name5']].astype('object')
    
    flag_list = []
    for i in range(len(dataSR_need)):
        flag = 0
        for j in range(1,5):
            name1 = dataSR_need.loc[i,'Name'+str(j)].upper() if pd.notnull(dataSR_need.loc[i,'Name'+str(j)]) else np.nan
            cost1 = dataSR_need.loc[i,'Cost'+str(j)]
            cost2 = dataSR_need.loc[i,'Cost'+str(j+1)]
            if name1 in approvers and pd.notnull(cost1) and pd.isnull(cost2):
                flag = 1
        flag_list.append(flag)
        
    dataSR_need['Flag_approver']=flag_list
    
    dataSR_need = dataSR_need[dataSR_need['Flag_approver']==1].reset_index().drop(['index'], axis = 1)

    data_need = pd.merge(dataMR_need, dataSR_need, on = 'SPR NO.', how = 'left')

    # filter CI
    def up(data, to_up):
        ctg = np.asarray(data[to_up])
        for i in range(0,len(ctg)):
            ctg[i] = ctg[i].upper()
        ctg_up = pd.DataFrame(ctg, columns = [to_up]) 
        data[to_up] = ctg_up[to_up]
        return data
    data_need['Category'] = data_need['Category'].astype(str)
    data = up(data_need, 'Category')
    data = data[data['Category'].str.contains('CI')]
    data = data[-(data['Category'].str.contains('CI Demo'))]
        
    # sales region
    data['Sales Region'] = np.where(data['Sales Region'].isnull()==True,'Others', data['Sales Region'])
    
    # SPR 
    data = data[(data['SPR NO.'].str.contains('RNE'))|(data['SPR NO.'].str.contains('RN'))|
                (data['SPR NO.'].str.contains('RS'))|(data['SPR NO.'].str.contains('RW'))|
                (data['SPR NO.'].str.contains('RE'))|(data['SPR NO.'].str.contains('RC'))|
                (data['SPR NO.'].str.contains('VS'))|(data['SPR NO.'].str.contains('OTH'))]
    data = data.reset_index().drop(['index'], axis = 1)

    now_line = get_cur_info()
    change_schedule(now_line,all_line)
    
    # list price fixed
    fixed_data.rename(columns = {'MLFB_with_Options':'MLFB'}, inplace = True)
    data = pd.merge(data, fixed_data, on = 'MLFB', how = 'left')
    data['L2_SAP'] = np.where(np.isnan(data['L2_SAP'])==True, data['List Price'], data['L2_SAP'])
    data.rename(columns = {'L2_SAP':'List Price Fixed'}, inplace = True)

    # cost
    cost = data[['SPR NO.', 'Price Group','Gross Margin', 'Gross Margin After SGA','Total Volume After Special Discount Off','Total Volume After Standard Discount Off']]
    cost['normal_cost'] = cost['Total Volume After Special Discount Off'] - cost['Gross Margin']*cost['Total Volume After Special Discount Off']
    cost['SGA_cost'] = cost['Total Volume After Special Discount Off'] - cost['Gross Margin After SGA']*cost['Total Volume After Special Discount Off']
    cost['final_cost'] = np.where(cost['SGA_cost'].isnull()==True,cost['normal_cost'], cost['SGA_cost'])
    
    # total cost&volume
    total_cost = cost[['SPR NO.', 'Price Group', 'final_cost', 'Total Volume After Standard Discount Off']]
    total_cost = total_cost.groupby(['SPR NO.','Price Group'], as_index = False).sum()  
    
    # margin
    total_cost['Margin_bf_spd (SAP cost)'] = (total_cost['Total Volume After Standard Discount Off']-total_cost['final_cost'])/total_cost['Total Volume After Standard Discount Off']
    total_cost = total_cost.drop(['Total Volume After Standard Discount Off'], axis = 1)

    # Total List Price Fixed
    total_price_fixed = data[['SPR NO.','Price Group', 'List Price Fixed', 'Agreed Qty']]
    total_price_fixed['PriceFixedxQty'] = total_price_fixed['List Price Fixed']*total_price_fixed['Agreed Qty']
    total_price_fixed = total_price_fixed.drop(["List Price Fixed", "Agreed Qty"], axis = 1).groupby(['SPR NO.','Price Group'], as_index = False).sum()  
    
    # total list price
    total_price = data[['SPR NO.', 'Price Group', 'List Price', 'Agreed Qty']]
    total_price['total_list_price'] = total_price['List Price']*total_price['Agreed Qty']
    total_price = total_price.drop(["List Price", "Agreed Qty"], axis = 1).groupby(['SPR NO.','Price Group'], as_index = False).sum()
    
    # list price change
    lp_change = total_price_fixed
    lp_change['List_Price_Change'] = total_price['total_list_price']/total_price_fixed['PriceFixedxQty']
    lp_change = lp_change.drop(['PriceFixedxQty'], axis = 1)
    
    # special discount
    discount = data[['SPR NO.', 'Price Group', 'Total Volume After Special Discount Off', 'Total Volume After Standard Discount Off']]
    discount['Requested Discount Off'] = (discount['Total Volume After Standard Discount Off']-discount['Total Volume After Special Discount Off'])/discount['Total Volume After Standard Discount Off']
    discount = discount.drop(['Total Volume After Special Discount Off'], axis = 1)
    
    discount = discount.merge(total_price[['SPR NO.','Price Group','total_list_price']],
                   on = ['SPR NO.','Price Group'], how = 'left')
    
    discount['Standard Discount Off'] = (discount['total_list_price']-discount['Total Volume After Standard Discount Off'])/discount['total_list_price']
    discount = discount.drop(['total_list_price'], axis = 1)
    
    # SPR_Total_List_Price
    data.rename(columns = {'Total Volume After Standard Discount Off':'Total Volume After Standard Discount Off ori'},inplace = True)
    spr = data[['SPR NO.','Total Volume After Standard Discount Off ori']].groupby('SPR NO.', as_index = False).sum()
    spr.rename(columns={'Total Volume After Standard Discount Off ori':'SPR_Total_List_Price'},inplace = True)
    
    data = pd.merge(data, spr, on = 'SPR NO.', how = 'left')
    data = pd.merge(data, discount, on = ['SPR NO.','Price Group'], how = 'left')
    data = pd.merge(data, total_price, on = ['SPR NO.','Price Group'], how = 'left')
    data = pd.merge(data, total_cost, on = ['SPR NO.','Price Group'], how = 'left')
    data = pd.merge(data, lp_change, on = ['SPR NO.','Price Group'], how = 'left')
        
    # Submitted Date
    data['Submitted Date'] = data['Submitted Date'].astype('datetime64').map(lambda x: x.strftime('%Y-%m-%d'))

    # customer value 
    data.rename(columns = {'End User NO.':'CNOC', 'End User Name(CN)':'Customer Name'}, inplace = True)
    #data['CNOC'] = data['CNOC'].astype('object')
    #c_value['CNOC'] = c_value['CNOC'].astype('object')
    cv_ci = c_value[c_value['segment_ADBA']=='CI']
    cv_pa = c_value.groupby('CNOC', as_index = False).sum()
    #cv_pa['CNOC'] = cv_pa['CNOC'].astype('object')
    data = pd.merge(data, cv_ci, on = 'CNOC', how = 'left').drop('segment_ADBA', axis = 1)
    data['sum_PurchaseVolume_201516'] = np.where(np.isnan(data['sum_PurchaseVolume_201516'])==True, 0, data['sum_PurchaseVolume_201516'])
    data['sum_PurchaseVolume_201617'] = np.where(np.isnan(data['sum_PurchaseVolume_201617'])==True, 0, data['sum_PurchaseVolume_201617'])
    data['sum_PurchaseVolume_201718'] = np.where(np.isnan(data['sum_PurchaseVolume_201718'])==True, 0, data['sum_PurchaseVolume_201718'])
    data['sum_PurchaseVolume_201819'] = np.where(np.isnan(data['sum_PurchaseVolume_201819'])==True, 0, data['sum_PurchaseVolume_201819'])
    
    # customer value ci
    data['Customer_Value_PY_CI_PA'] = 0
    data['Customer_Value_PY_CI_PA'] = np.where(data['Fiscal Year']=='16/17', data['sum_PurchaseVolume_201516'], data['Customer_Value_PY_CI_PA'])
    data['Customer_Value_PY_CI_PA'] = np.where(data['Fiscal Year']=='17/18', data['sum_PurchaseVolume_201617'], data['Customer_Value_PY_CI_PA'])
    data['Customer_Value_PY_CI_PA'] = np.where(data['Fiscal Year']=='18/19', data['sum_PurchaseVolume_201718'], data['Customer_Value_PY_CI_PA'])
    data['Customer_Value_PY_CI_PA'] = np.where(data['Fiscal Year']=='19/20', data['sum_PurchaseVolume_201819'], data['Customer_Value_PY_CI_PA'])
    
    # ci new
    data['CI_New'] = np.where((data['sum_PurchaseVolume_201516']==0)&(data['sum_PurchaseVolume_201617']==0)&(data['sum_PurchaseVolume_201718']==0)&(data['sum_PurchaseVolume_201819']==0),1,0)
    data = data.drop(['sum_PurchaseVolume_201516', 'sum_PurchaseVolume_201617',
                      'sum_PurchaseVolume_201718', 'sum_PurchaseVolume_201819',],axis = 1)
    
    # customer value pa
    data = pd.merge(data, cv_pa, on = 'CNOC', how = 'left')
    data['sum_PurchaseVolume_201516'] = np.where(np.isnan(data['sum_PurchaseVolume_201516'])==True, 0, data['sum_PurchaseVolume_201516'])
    data['sum_PurchaseVolume_201617'] = np.where(np.isnan(data['sum_PurchaseVolume_201617'])==True, 0, data['sum_PurchaseVolume_201617'])
    data['sum_PurchaseVolume_201718'] = np.where(np.isnan(data['sum_PurchaseVolume_201718'])==True, 0, data['sum_PurchaseVolume_201718'])
    data['sum_PurchaseVolume_201819'] = np.where(np.isnan(data['sum_PurchaseVolume_201819'])==True, 0, data['sum_PurchaseVolume_201819'])
    data['Customer_Value_PY_PA'] = 0
    data['Customer_Value_PY_PA'] = np.where(data['Fiscal Year']=='16/17', data['sum_PurchaseVolume_201516'], data['Customer_Value_PY_PA'])
    data['Customer_Value_PY_PA'] = np.where(data['Fiscal Year']=='17/18', data['sum_PurchaseVolume_201617'], data['Customer_Value_PY_PA'])
    data['Customer_Value_PY_PA'] = np.where(data['Fiscal Year']=='18/19', data['sum_PurchaseVolume_201718'], data['Customer_Value_PY_PA'])
    data['Customer_Value_PY_PA'] = np.where(data['Fiscal Year']=='19/20', data['sum_PurchaseVolume_201819'], data['Customer_Value_PY_PA'])
    data = data.drop(['sum_PurchaseVolume_201516', 'sum_PurchaseVolume_201617',
                      'sum_PurchaseVolume_201718', 'sum_PurchaseVolume_201819',], axis = 1)

    now_line = get_cur_info()
    change_schedule(now_line,all_line)
    
    # customer branch
    c_branch['CNOC'] = c_branch['CNOC'].astype('object')
    data = pd.merge(data, c_branch, on = 'CNOC', how = 'left')
    data['Industry'] = np.where(data['Industry'].isnull()==True,'Others', data['Industry'])
    data['Industry'] = np.where((data['Industry'].str.contains('地铁'))|(data['Industry'].str.contains('轨道交通')),'地铁/轨道交通',data['Industry'] )
    
    # branch
    branches = ['Automotive','Food & Beverages','Glass & Solar','Hospitality & Entertainment',
                'Metals','Power Utilities','Public Sector','Transportation',
                'Wholesale/Retail','地铁/轨道交通','Chemicals',
                'Electrical & Electronic (E&E)','Machine Building',
                'Minerals','Others','Post & Logistics','Construction/Real Estate',
                'Marine','Oil & Gas','Water']
    for branch in branches:
        data[branch] = np.where((data['Industry']==branch),1,0)
    
    # project oem partner
    data['Project'] = np.where((data['Category'].str.contains('PROJECT')),1,0)
    data['OEM'] = np.where((data['Category'].str.contains('OEM')),1,0)
    data['Partner'] = np.where((data['Category'].str.contains('PARTNER')),1,0)
    
    # Is_Yearly
    data['Is_Yearly'] = np.where((data['SPR NO.'].str.contains('Y')),1,0)   

    # cmm
    data['CMM'] = np.where((data['Submitted Date']>='2017-10-01')&(data['Submitted Date']<='2018-09-30'), 3957, 0)
    data['CMM'] = np.where((data['Submitted Date']>='2018-10-01')&(data['Submitted Date']<='2019-09-30'), 4036, data['CMM'])
    data['CMM'] = np.where((data['Submitted Date']>='2019-10-01')&(data['Submitted Date']<='2020-09-30'), 4121, data['CMM'])

    # PG_Count
    pg_count = data[['SPR NO.', 'Price Group']].drop_duplicates().reset_index().drop(['index'], axis = 1).groupby('SPR NO.', as_index = False).count()
    pg_count.rename(columns = {'Price Group':'PG_Count'}, inplace = True)
    data = pd.merge(data, pg_count, on = 'SPR NO.', how = 'left') 
    
    # date
    data['year_split'] = data['Submitted Date'].astype('datetime64').dt.year
    data['month_split'] = data['Submitted Date'].astype('datetime64').dt.month
    data['month_temp'] = data['Submitted Date'].astype('datetime64').dt.month
    data['month_split'] = np.where(data['month_temp']!=1,data['month_split']-1,12)
    data['year_split'] = np.where(data['month_temp']==1,data['year_split']-1,data['year_split'] )
    data['Month'] = pd.to_datetime(data['year_split'].astype(str) +'-'+ data['month_split'].astype(str), format='%Y-%m')
    data['Month'] = data['Month'].map(lambda x: x.strftime('%Y-%m'))
    data = data.drop(['year_split', 'month_split', 'month_temp'], axis = 1)
    
    # index
    index_data.rename(columns = {'Year_Month':'Month', 
                                 'PriceGroup':'Price Group'}, inplace = True)
    index_data['Price Group'] = index_data['Price Group'].astype(str)
    data['Price Group'] = data['Price Group'].astype(str)
    data = pd.merge(data, index_data, on = ['Month', 'Price Group'], how = 'left')
    # index extra
    index_data_extra.rename(columns = {'Year_Month':'Month','PriceGroup_index':'PriceGroup_index_CI'}, inplace = True)
    data = pd.merge(data, index_data_extra, on = 'Month', how = 'left')
    data['PriceGroup_index'] = np.where(data['PriceGroup_index'].isnull()==True, data['PriceGroup_index_CI'], data['PriceGroup_index'])
    data = data.drop(['PriceGroup_index_CI'], axis = 1)

    now_line = get_cur_info()
    change_schedule(now_line,all_line)

    # independent
    col_all = ['SPR NO.','Price Group','List_Price_Change',
               'Standard Discount Off','SPR_Total_Price_after_std',
               'PG_Count','Customer_Value_PY_PA','Customer_Value_PY_CI_PA','Margin_bf_spd (SAP cost)',
               'CI_New','Is_Yearly','PriceGroup_index','CMM','Food & Beverages','Glass & Solar',
               'Hospitality & Entertainment','Metals','Power Utilities','Public Sector',
               'Transportation','Wholesale/Retail','地铁/轨道交通','OEM','Partner','Project',
               'Automotive','Chemicals','Electrical & Electronic (E&E)','Machine Building',
               'Minerals','Others','Post & Logistics','Construction/Real Estate',
               'Marine','Oil & Gas','Water']
    data_input = data[col_all].drop_duplicates().reset_index().drop(['index'], axis = 1)
    data_input['List_Price_Change'] = np.where(data_input['List_Price_Change'].isnull()==True, 1, data_input['List_Price_Change'])
    data_input['Customer_Value_PY_PA'] = np.where(data_input['Customer_Value_PY_PA'].isnull()==True, 0, data_input['Customer_Value_PY_PA'])
    data_input['Customer_Value_PY_CI_PA'] = np.where(data_input['Customer_Value_PY_CI_PA'].isnull()==True, 0, data_input['Customer_Value_PY_CI_PA'])
    
    nas = ['Standard Discount Off','SPR_Total_Price_after_std',
           'PG_Count','Margin_bf_spd (SAP cost)',
           'Is_Yearly','PriceGroup_index']
    data_error = pd.DataFrame()
    for na in nas:
        temp = data_input[data_input[na].isnull()]
        data_error = pd.concat([data_error, temp],axis=0)
    data_error = data_error[['SPR NO.', 'Price Group']].drop_duplicates().reset_index().drop(['index'], axis = 1)

    if data_error.empty==False:
        missing_output = os.path.dirname(filename) + '\\error log ' + time.strftime("%Y%m%d") + '.txt'
        data_error.to_csv(missing_output ,header = True, index = False)
    
    for na in nas:
        data_input[na] = np.where(data_input[na].isnull()==True,0,data_input[na])

    data_final_bfstd = copy.deepcopy(data)

    col_sc = ['List_Price_Change','Standard Discount Off','SPR_Total_Price_after_std',
              'PG_Count','Customer_Value_PY_PA','Customer_Value_PY_CI_PA','Margin_bf_spd (SAP cost)',
              'CI_New','Is_Yearly','PriceGroup_index','CMM']
    data2std = data_input[col_sc]
    data2std['Margin_bf_spd (SAP cost)'] = data2std['Margin_bf_spd (SAP cost)']*100
    data2std['Standard Discount Off'] = data2std['Standard Discount Off']*100
    
    try:
        data_std = sc.transform(data2std)
    except Exception:
        print(traceback.format_exc())
        error_output = os.path.dirname(filename) + '\\error log ' + time.strftime("%Y%m%d") + '.txt'
        with open(error_output, 'w+') as fff:
            fff.write(str(traceback.format_exc()))
        mb.showerror(title="Error", message='Error! Saved to ' + error_output)  
        root.destroy()
        
    data_std = pd.DataFrame(data_std, columns = col_sc)
  
    col_mean = ['Food & Beverages','Glass & Solar','Hospitality & Entertainment',
                'Metals','Power Utilities','Public Sector','Transportation',
                'Wholesale/Retail','地铁/轨道交通','Standard Discount Off',
                'SPR_Total_Price_after_std','PG_Count',
                'Customer_Value_PY_PA','Customer_Value_PY_CI_PA',
                'Margin_bf_spd (SAP cost)','Is_Yearly','CMM']
    mean_sc = ['Standard Discount Off','SPR_Total_Price_after_std',
                'PG_Count','Customer_Value_PY_PA','Customer_Value_PY_CI_PA',
                'Margin_bf_spd (SAP cost)','Is_Yearly','CMM']
    independent_mean = data_input[col_mean]
    independent_mean[mean_sc] = data_std[mean_sc]

    now_line = get_cur_info()
    change_schedule(now_line,all_line)
     
    col_80 = ['OEM','Partner','Project','Automotive','Chemicals',
              'Electrical & Electronic (E&E)','Food & Beverages',
              'Hospitality & Entertainment','Machine Building',
              'Metals','Minerals','Others','Post & Logistics',
              'Power Utilities','Public Sector','Transportation',
              'Wholesale/Retail','地铁/轨道交通','List_Price_Change',
              'Standard Discount Off','SPR_Total_Price_after_std',
              'PG_Count', 'Customer_Value_PY_PA','Customer_Value_PY_CI_PA',
              'Margin_bf_spd (SAP cost)','CI_New','Is_Yearly',
              'PriceGroup_index','CMM']
    upper80_sc = [ 'List_Price_Change','Standard Discount Off',
              'SPR_Total_Price_after_std','PG_Count',
              'Customer_Value_PY_PA', 'Customer_Value_PY_CI_PA',
              'Margin_bf_spd (SAP cost)', 'CI_New',
              'Is_Yearly', 'PriceGroup_index','CMM']
    independent80 = data_input[col_80]
    independent80[upper80_sc] = data_std[upper80_sc]

    col_95 = ['OEM','Partner','Project','Automotive','Chemicals','Construction/Real Estate',
              'Electrical & Electronic (E&E)','Food & Beverages','Machine Building',
              'Marine','Metals','Minerals', 'Oil & Gas','Others',
              'Post & Logistics','Power Utilities','Public Sector','Transportation',
              'Water','Wholesale/Retail','地铁/轨道交通','List_Price_Change',
              'Standard Discount Off','SPR_Total_Price_after_std',
              'PG_Count','Customer_Value_PY_PA','Customer_Value_PY_CI_PA',
              'Margin_bf_spd (SAP cost)','CI_New','PriceGroup_index','CMM']
    upper95_sc = ['List_Price_Change','Standard Discount Off',
              'SPR_Total_Price_after_std','PG_Count',
              'Customer_Value_PY_PA','Customer_Value_PY_CI_PA',
              'Margin_bf_spd (SAP cost)','CI_New','PriceGroup_index','CMM']
    independent95 = data_input[col_95]
    independent95[upper95_sc] = data_std[upper95_sc]

    now_line = get_cur_info()
    change_schedule(now_line,all_line)
    
    # predict
    # result = data_input[['SPR NO.','Price Group']]
    independent4model = independent_mean.astype(float)
    independent4model_value = xgb.DMatrix(independent4model)
    result_temp = mean.predict(independent4model_value)
    result = pd.DataFrame(result_temp, columns = ['Adviced special discount off'])
    result_temp_up = upper80.predict(independent80)
    result_up = pd.DataFrame(result_temp_up, columns = ['80% interval (max)'])
    result_temp_up95 = upper95.predict(independent95)
    result_up95 = pd.DataFrame(result_temp_up95, columns = ['95% interval (max)'])
    result['80% interval (max)'] = result_up['80% interval (max)']    
    result['95% interval (max)'] = result_up95['95% interval (max)']  
    result[['SPR NO.','Price Group']] = data_input[['SPR NO.','Price Group']]
    data_final_bfstd = pd.merge(data_final_bfstd, result, on = ['SPR NO.','Price Group'], how = 'left')
    
    # output data
    data_final_bfstd['95% interval (max)'] = np.where(data_final_bfstd['95% interval (max)']<data_final_bfstd['80% interval (max)'],data_final_bfstd['80% interval (max)'],data_final_bfstd['95% interval (max)'])    
    data_final_bfstd['Light'] = np.where(data_final_bfstd['Requested Discount Off']>data_final_bfstd['95% interval (max)'],1,0)
    data_final_bfstd['Light'] = np.where((data_final_bfstd['Requested Discount Off']>data_final_bfstd['80% interval (max)'])&(data_final_bfstd['Requested Discount Off']<=data_final_bfstd['95% interval (max)']),2,data_final_bfstd['Light'])
    data_final_bfstd['Light'] = np.where(data_final_bfstd['Requested Discount Off']<=data_final_bfstd['80% interval (max)'],3,data_final_bfstd['Light'])
    data_final_bfstd['95% interval (max)'] = data_final_bfstd['95% interval (max)'].apply(lambda x: format(x, '.1%'))
    data_final_bfstd['80% interval (max)'] = data_final_bfstd['80% interval (max)'].apply(lambda x: format(x, '.1%'))
    data_final_bfstd['Requested Discount Off'] = data_final_bfstd['Requested Discount Off'].apply(lambda x: format(x, '.1%'))    
    data_final_bfstd['Adviced special discount off'] = data_final_bfstd['Adviced special discount off'].apply(lambda x: format(x, '.1%'))
    data_final_bfstd['Standard Discount Off'] = data_final_bfstd['Standard Discount Off'].apply(lambda x: format(x, '.1%'))    
    data_final_bfstd['Customer_Value_PY_PA'] = round(data_final_bfstd['Customer_Value_PY_PA'])
    data_final_bfstd['Customer_Value_PY_CI_PA'] = round(data_final_bfstd['Customer_Value_PY_CI_PA'])    
    data_final_bfstd['Is_Yearly'] = np.where(data_final_bfstd['Is_Yearly']==1,'Y','N')
    
    now_line = get_cur_info()
    change_schedule(now_line,all_line)
    #data_pg = pd.DataFrame([],columns = col_keep_pg)list(data_pg)
    # by price group
    col_keep_pg = ['SPR NO.','Price Group','Requested Discount Off','Light',
                   '80% interval (max)','95% interval (max)',
                   'Submitted Date','Category','Is_Yearly','CNOC','Customer Name','Industry',
                   'total_list_price','final_cost','Standard Discount Off',
                   'Margin_bf_spd (SAP cost)','PriceGroup_index',
                   'PG_Count','Total Volume After Standard Discount Off',
                   'Customer_Value_PY_PA',
                   'Customer_Value_PY_CI_PA','CI_New','CMM',
                   'Name1','Cost1','Name2','Cost2','Name3',
                   'Cost3','Name4','Cost4','Name5','Cost5'] 
    data_pg = data_final_bfstd[col_keep_pg].drop_duplicates().reset_index().drop(['index'], axis = 1)
    data_pg.rename(columns = {'total_list_price':'Total List Pice','final_cost':'Total Purchase',
                              'Is_Yearly':'Yearly SPR','Margin_bf_spd (SAP cost)':'Margin after STD',
                              'PG_Count':'PG count','PriceGroup_index':'Price index (by PG)',
                              'Industry':'Branch in TrustIT','Customer_Value_PY_PA':'Customer value PY (PA)',
                              'Customer_Value_PY_CI_PA':'Customer value PY (PA CI)',
                              'CNOC':'End User NO.', 'Customer Name':'End User Name(CN)',
                              'Light':'Alert',
                              '80% interval (max)':'80% safety level',
                              '95% interval (max)':'95% safety level'}, inplace = True)
    data_pg['Total List Pice'] = round(data_pg['Total List Pice'])
    data_pg['Total Purchase'] = round(data_pg['Total Purchase'])
    data_pg['Margin after STD'] = data_pg['Margin after STD'].apply(lambda x: format(x, '.1%'))
    for i in range(len(data_pg['Price index (by PG)'])):
        data_pg['Price index (by PG)'][i] = '%.2f' % data_pg['Price index (by PG)'][i]
    # file_output_bfstd = 'Predict.xlsx'
    data_pg = data_pg.sort_values(['Alert','SPR NO.','Price Group'],ascending=True)
    data_pg.to_excel(file_output_bfstd, header = True, index = False, sheet_name = 'by Price Group')

    now_line = get_cur_info()
    change_schedule(now_line,all_line)

    # xlwings
    pythoncom.CoInitialize()
    app = xw.App(visible = False, add_book = False)
    wb = app.books.open(file_output_bfstd) #打开文件
    ws_pg = wb.sheets['by Price Group']
    
    #last_column_pg = ws_pg.range(1, 1).end('right').get_address(0, 0)[0] #获取最后列
    last_column_pg = 'AG'
    last_row_pg = ws_pg.range(1, 1).end('down').row #获取最后行
    a_range_pg = f'A1:{last_column_pg}{last_row_pg}' #生成表格的数据范围
    ws_pg.range(a_range_pg).api.Borders(8).LineStyle = 1 #上边框
    ws_pg.range(a_range_pg).api.Borders(9).LineStyle = 1 #下边框
    ws_pg.range(a_range_pg).api.Borders(7).LineStyle = 1 #左边框
    ws_pg.range(a_range_pg).api.Borders(10).LineStyle = 1 #右边框
    ws_pg.range(a_range_pg).api.Borders(12).LineStyle = 1 #内横边框
    ws_pg.range(a_range_pg).api.Borders(11).LineStyle = 1 #内纵边框  
    ws_pg.range(a_range_pg).columns.autofit()

    wb.save(file_output_bfstd)
    wb.close()
    app.quit()

    now_line = get_cur_info()
    change_schedule(now_line,all_line)
    
    # openpyxl    
    excel = load_workbook(file_output_bfstd)
    table = excel['by Price Group']     
    rows = table.max_row   
    ws = excel.active
    rule = IconSetRule('3TrafficLights1', 'num', [1, 2, 3], showValue=False, percent=None, reverse=None)
    ws.conditional_formatting.add(f'D2:D{rows}', rule)
    for row in range(2, rows+1):
        light = ws.cell(row=row, column=4)
        light.alignment = Alignment(horizontal='center', vertical='center')

    excel.save(file_output_bfstd)

    now_line = all_line
    change_schedule(now_line,all_line)
   
    if data_error.empty==False:
        mes = 'Finished with error(s)! Please see error log for details.'
        mb.showinfo(message = mes)
    else:
        mes = 'Finished!'
        mb.showinfo(message = mes)
    root.destroy()

root = tk.Tk()
width = 450
height = 120
screenwidth = root.winfo_screenwidth()
screenheight = root.winfo_screenheight()
alignstr = '%dx%d+%d+%d' % (width, height, (screenwidth-width)/2, ((screenheight-height)/2)-0.2*screenheight)
root.geometry(alignstr)

# 设置窗口大小不可改变
root.resizable(width=False,height=False)
#root.geometry('450x120+100+20')
root.title('SPR Price Check - PA CI')

# 画布        
canvas = tk.Canvas(root, bg = 'whitesmoke')
canvas.place(relwidth = 1, relheight = 1)

# 选择文件的框
frame = tk.Frame(root, bg='#D9D9D9')
frame.place(relx=0.01,rely=0.05, relwidth=0.98,relheight=0.85)

# 下载速度的框
entbar = tk.Canvas(root, bg='whitesmoke')
entbar.place(relx=0,rely=0.94, relwidth=1,relheight=0.06)
# 进度条
fill_rec = entbar.create_rectangle(0,0,0,36,outline = "",width = 0,fill = "#08A752")
                                   
#begin = threading.Thread(target=run)
#begin.setDaemon(True)
#begin.start()
  
# 文字
textt = 'Processing...\n\n' + filename
labname = tk.Label(frame,text = textt, justify='left', foreground = 'black', 
                   background='#D9D9D9', width=150, height = 10, font = ('arial', 9),wraplength = 400)
labname.AutoSize = True
labname.place(relx = 0.01, rely = 0, relwidth = 0.95, relheight = 0.8)

ext=tk.Button(root,width=10,text='Run', justify='right', background='#1D87AF', foreground = 'white', 
              command=run, font = ('arial', 10, 'bold'), relief = tk.RAISED)
ext.place(relx=0.41,rely=0.68, relwidth=0.18,relheight=0.2)
                                 
root.mainloop()
